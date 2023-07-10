import time, os, datetime, sys, glob
import openpyxl as xl;

start_time = time.time()
# Today's date to open the correct file
todayDate = ' ' + datetime.date.today().strftime("%d%m%Y")
print('Today is:' + todayDate)

source_dir = 'XXXXXX'
dest_dir = os.path.normpath(os.path.expanduser("~/Desktop"))

# [HW]-Open Stock HUAWEI TODAY:
listing = glob.glob(dest_dir + '/Stock Spain*')
for filename in listing:
    stockHW = filename
print("[HW]->>> Opening Excel file: " + XXXXXX + " in Desktop...")

try:
    wb1 = xl.load_workbook(XXXXXX)
except FileNotFoundError:
    print(f'[HW]-The file: {XXXXXX} does not exist in DESKTOP')
    sys.exit()

ws1 = wb1.worksheets[0]

# [HW]-Creating the destination excel file
print("[HW]-Creating target file...") 
filename1 = dest_dir + "/" + "Stock Spain " + todayDate + " PYTHONtoSendTo_FTP.xlsx"
sheetNameforImport = 'Sheet1'
wb2=xl.Workbook()
ws2 = wb2.active
ws2.title = sheetNameforImport
ws2.title = sheetNameforImport

# [HW]-The items types to KEEP
my_file = open(r'D:\Python\New FILES\Filtros\Filter_Warranty_HW.txt', 'r')
#read text file into list 
ITEMStoKEEP = my_file.read().split('\n')

# [HW]-Calculate total number of rows and columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

mrPrint = f'{mr:,}'.replace(',','.') #Convert to XXX.XXX
print("[HW]-Maximum rows before removing:", mrPrint)
print("[HW]-Removing lines...")

# [XXXXXX]-copying the cell values from source excel file to destination excel file
DestIntRow = 1
for i in range (1, mr + 1):
    #Only Copy files in ITEMStoKEEP list 
    cellcheck = ws1.cell(row = i, column = 8)
    if cellcheck.value in ITEMStoKEEP:
        for j in range (1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row = i, column = j)
            # writing the read value to destination excel file
            ws2.cell(row = DestIntRow, column = j).value = c.value
        DestIntRow = DestIntRow + 1 
        
# [XXXXXX]-destination final row
Dmr = ws2.max_row
DmrPrint = f'{Dmr:,}'.replace(',','.') #Convert to XXX.XXX
print("[HW]-Maximum rows in file to import for OWS:",DmrPrint)

# [XXXXXX]-changing sheet name & saving the destination excel file
ws2.title = sheetNameforImport
wb2.save(str(filename1))

print(f'[XXXXXX]-Done! Completed in {round(time.time()-start_time,2)} seconds.')
laptime = round(time.time(),2)
print(f'----------------------------------------------------------------------------------------')

print(f'>>Upload File FTP [AUTO]: FTPsendFile')
import FTPsendFile


print(f'[TOTAL]-Done! Completed in {round(time.time()-start_time,2)} seconds.')
